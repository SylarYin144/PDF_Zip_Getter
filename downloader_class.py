import pandas as pd
import requests
import zipfile
import os
import re
import time
import sys
from datetime import datetime
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import json
import xml.etree.ElementTree as ET
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import base64
import threading
import queue
import itertools
import openpyxl

# --- Configuration Constants ---
STANDARD_USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36'

class Downloader:
    def __init__(self, config, progress_queue, cancel_event, pause_event):
        self.config = config
        self.progress_queue = progress_queue
        self.cancel_event = cancel_event
        self.pause_event = pause_event
        self.driver = None
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': STANDARD_USER_AGENT})
        self.temp_pdf_paths = []

    # --- Communication Methods ---
    def _send_log(self, message):
        self.progress_queue.put({'type': 'log', 'message': message})

    def _send_progress(self, current, total, doi, status):
        self.progress_queue.put({
            'type': 'progress',
            'current': current,
            'total': total,
            'doi': doi,
            'status': status
        })

    def _send_article_result(self, doi, success, source=None, reason=None):
        self.progress_queue.put({
            'type': 'article_result',
            'doi': doi,
            'success': success,
            'source': source,
            'reason': reason,
        })

    def _send_kpi_update(self, obtained, failed, pending):
        self.progress_queue.put({
            'type': 'kpi',
            'obtained': obtained,
            'failed': failed,
            'pending': pending
        })

    # --- Core Logic ---
    def run(self):
        self._send_log("Proceso de descarga iniciado.")

        if self.config.get('use_google_scholar') or self.config.get('use_pmc'):
            if not self._initialize_driver():
                self._send_log("ADVERTENCIA: No se pudo iniciar Selenium. Las fuentes Google Scholar y PMC no estarán disponibles.")

        if self.config.get('excel_report_path'):
            self._create_initial_report()

        df = self.config['input_df']
        total_articles = len(df)
        zip_path = self.config['zip_path']

        successful_articles_data = []
        failed_articles_data = []
        was_cancelled = False

        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for index, row in df.iterrows():
                    try:
                        self._check_pause_cancel()
                    except InterruptedError:
                        was_cancelled = True
                        break

                    original_row_data = row.to_dict()
                    doi = str(original_row_data.get('DOI', '')).strip()
                    title = str(original_row_data.get('Title', '')).strip()
                    effective_title = title if title else doi

                    self._send_progress(index + 1, total_articles, doi, "Iniciando...")
                    self._send_kpi_update(len(successful_articles_data), len(failed_articles_data), total_articles - index)

                    if not doi:
                        self._send_log(f"DOI vacío en la fila {index+1}, saltando.")
                        reason = "DOI vacío"
                        self._send_article_result(doi, success=False, reason=reason)
                        failed_articles_data.append({'data': original_row_data, 'reason': reason})
                        if self.config.get('excel_report_path'): self._update_report_on_failure(original_row_data, reason)
                        continue

                    pdf_content, download_source, failure_reason = self._try_all_sources(doi, effective_title, index, total_articles)

                    if pdf_content:
                        successful_articles_data.append({'data': original_row_data, 'source': download_source})
                        pdf_filename_in_zip = self._clean_filename(effective_title)[:150] + ".pdf"

                        temp_dir = "temp_scihub_pdfs"
                        if not os.path.exists(temp_dir): os.makedirs(temp_dir)
                        temp_pdf_path = os.path.join(temp_dir, f"temp_{os.getpid()}_{index}.pdf")

                        with open(temp_pdf_path, 'wb') as f: f.write(pdf_content)
                        zf.write(temp_pdf_path, arcname=pdf_filename_in_zip)
                        self.temp_pdf_paths.append(temp_pdf_path)

                        self._send_article_result(doi, success=True, source=download_source)
                        if self.config.get('excel_report_path'):
                            self._update_report_on_success(original_row_data, download_source)
                    else:
                        failed_articles_data.append({'data': original_row_data, 'reason': failure_reason})
                        self._send_article_result(doi, success=False, reason=failure_reason)
                        if self.config.get('excel_report_path'):
                            self._update_report_on_failure(original_row_data, failure_reason)

                    self._send_kpi_update(len(successful_articles_data), len(failed_articles_data), total_articles - (index + 1))

                    time.sleep(self.config['inter_doi_delay'])

        except InterruptedError:
            was_cancelled = True
        except Exception as e:
            self._send_log(f"Error crítico durante el proceso de descarga: {e}")

        finally:
            self._cleanup()
            results = {
                'total_articles': total_articles,
                'successful_count': len(successful_articles_data),
                'failed_count': len(failed_articles_data),
                'successful_articles': successful_articles_data,
                'failed_articles': failed_articles_data,
                'was_cancelled': was_cancelled,
                'zip_path': self.config['zip_path'],
                'excel_report_path': self.config.get('excel_report_path')
            }
            self.progress_queue.put({'type': 'finished', 'results': results})

    def _try_all_sources(self, doi, effective_title, index, total_articles):
        # 1. Sci-Hub
        self._send_progress(index + 1, total_articles, doi, "Buscando en Sci-Hub...")
        for mirror in self.config['sci_hub_mirrors']:
            self._check_pause_cancel()
            if self.cancel_event.is_set(): return None, None, "Cancelado"

            self._send_log(f"Intentando con mirror de Sci-Hub: {mirror}")
            pdf_content, reason = self._download_from_scihub_mirror(mirror, doi)
            if pdf_content:
                self._send_log(f"Éxito con Sci-Hub mirror: {mirror}")
                return pdf_content, "Sci-Hub", None
            else:
                self._send_log(f"Fallo con Sci-Hub mirror {mirror}: {reason}")

        # 2. Google Scholar
        if self.config['use_google_scholar'] and self.driver:
            self._check_pause_cancel()
            if self.cancel_event.is_set(): return None, None, "Cancelado"
            self._send_progress(index + 1, total_articles, doi, "Buscando en Google Scholar...")
            pdf_content, reason = self._download_with_selenium_google_scholar(doi, effective_title)
            if pdf_content:
                self._send_log(f"Éxito con Google Scholar: {reason}")
                return pdf_content, "Google Scholar", None
            else:
                self._send_log(f"Fallo con Google Scholar: {reason}")

        # 3. PubMed Central (PMC)
        if self.config['use_pmc'] and self.driver:
            self._check_pause_cancel()
            if self.cancel_event.is_set(): return None, None, "Cancelado"
            self._send_progress(index + 1, total_articles, doi, "Buscando en PubMed Central...")
            pdf_content, reason = self._download_with_selenium_pmc(doi, effective_title)
            if pdf_content:
                self._send_log(f"Éxito con PubMed Central: {reason}")
                return pdf_content, "PubMed Central", None
            else:
                self._send_log(f"Fallo con PubMed Central: {reason}")

        return None, None, "No encontrado en ninguna fuente seleccionada"

    def _download_from_scihub_mirror(self, mirror_url, doi):
        full_url = f"{mirror_url}{doi}"
        try:
            response = self.session.get(full_url, timeout=30)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')

            iframe = soup.find('iframe', id='pdf')
            if iframe and iframe.get('src'):
                pdf_src = iframe['src']
                if pdf_src.startswith('//'): pdf_src = 'https:' + pdf_src
                elif pdf_src.startswith('/'): pdf_src = urljoin(full_url, pdf_src)

                pdf_response = self.session.get(pdf_src, timeout=60)
                pdf_response.raise_for_status()
                if 'application/pdf' in pdf_response.headers.get('Content-Type', '').lower():
                    return pdf_response.content, "Éxito (iframe)"

            if 'application/pdf' in response.headers.get('Content-Type', '').lower():
                return response.content, "Éxito (directo)"

            return None, "No se encontró enlace PDF en la página"
        except requests.exceptions.RequestException as e:
            return None, f"Error de red: {e}"
        except Exception as e:
            return None, f"Error inesperado: {e}"

    def _download_with_selenium_google_scholar(self, doi, title):
        wait_timeout = self.config.get('element_wait_timeout', 20)
        try:
            self._send_log(f"SELENIUM GS: Buscando DOI: {doi}")
            scholar_url = f"https://scholar.google.com/scholar?hl=en&q={doi}"
            self.driver.get(scholar_url)

            wait = WebDriverWait(self.driver, wait_timeout)
            pdf_links = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//a[contains(translate(., 'PDF', 'pdf'), 'pdf') or contains(@href, '.pdf')]")))

            for link_element in pdf_links:
                pdf_url = link_element.get_attribute('href')
                if pdf_url and 'scholar.google.com' not in pdf_url:
                    self._send_log(f"SELENIUM GS: Encontrado posible PDF en: {pdf_url}")
                    pdf_content = self._get_pdf_content_via_js(pdf_url)
                    if pdf_content:
                        return pdf_content, f"Obtenido de {pdf_url}"
            return None, "No se encontró un enlace PDF válido."
        except TimeoutException:
            return None, "Timeout esperando enlaces PDF en la página."
        except Exception as e:
            return None, f"Error en Selenium: {e}"

    def _download_with_selenium_pmc(self, doi, title):
        wait_timeout = self.config.get('element_wait_timeout', 20)
        try:
            self._send_log(f"SELENIUM PMC: Buscando DOI: {doi}")
            search_url = f"https://www.ncbi.nlm.nih.gov/pmc/?term={doi}"
            self.driver.get(search_url)
            wait = WebDriverWait(self.driver, wait_timeout)

            article_links = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.rprt .title a")))
            if not article_links: return None, "No se encontró el artículo en PMC."

            article_url = article_links[0].get_attribute('href')
            self.driver.get(article_url)

            pdf_links = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@href, '.pdf') and contains(., 'PDF')]")))
            if not pdf_links: return None, "No se encontró enlace PDF en la página del artículo."

            pdf_url = pdf_links[0].get_attribute('href')
            self._send_log(f"SELENIUM PMC: Encontrado posible PDF en: {pdf_url}")
            pdf_content = self._get_pdf_content_via_js(pdf_url)
            if pdf_content: return pdf_content, f"Obtenido de {pdf_url}"

            return None, "No se pudo descargar el PDF desde el enlace."
        except TimeoutException:
            return None, "Timeout esperando elementos en la página."
        except Exception as e:
            return None, f"Error en Selenium: {e}"

    def _get_pdf_content_via_js(self, pdf_url):
        script = """
        const callback = arguments[arguments.length - 1];
        fetch(arguments[0])
            .then(response => response.blob())
            .then(blob => {
                const reader = new FileReader();
                reader.onloadend = () => {
                    const base64Marker = ';base64,';
                    const base64Data = reader.result.substring(reader.result.indexOf(base64Marker) + base64Marker.length);
                    callback(base64Data);
                };
                reader.onerror = () => callback({error: 'FileReader error'});
                reader.readAsDataURL(blob);
            })
            .catch(error => callback({error: 'JS Fetch error: ' + error}));
        """
        try:
            self.driver.set_script_timeout(self.config.get('page_load_timeout', 60))
            result = self.driver.execute_async_script(script, pdf_url)
            if isinstance(result, dict) and 'error' in result:
                self._send_log(f"JS Fetch Helper Error: {result['error']}")
                return None
            return base64.b64decode(result) if result else None
        except Exception as e:
            self._send_log(f"JS Fetch Helper Exception: {e}")
            return None

    def _initialize_driver(self):
        try:
            self._send_log("Inicializando WebDriver de Selenium...")
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument(f"user-agent={STANDARD_USER_AGENT}")
            self.driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
            self.driver.set_page_load_timeout(self.config.get('page_load_timeout', 60))
            self._send_log("WebDriver de Selenium inicializado.")
            return True
        except Exception as e:
            self._send_log(f"Error al inicializar WebDriver: {e}")
            self.driver = None
            return False

    def _cleanup(self):
        self._send_log("Limpiando recursos...")
        if self.driver:
            try: self.driver.quit()
            except Exception as e: self._send_log(f"Error al cerrar WebDriver: {e}")
        for temp_path in self.temp_pdf_paths:
            try: os.remove(temp_path)
            except OSError: pass
        temp_dir = "temp_scihub_pdfs"
        if os.path.exists(temp_dir) and not os.listdir(temp_dir):
            try: os.rmdir(temp_dir)
            except OSError: pass
        self._send_log("Limpieza finalizada.")

    def _check_pause_cancel(self):
        while self.pause_event.is_set():
            if self.cancel_event.is_set(): break
            time.sleep(1)
        if self.cancel_event.is_set():
            self._send_log("Cancelación detectada. Terminando...")
            raise InterruptedError("Proceso cancelado por el usuario.")

    def _clean_filename(self, title):
        return re.sub(r'[\\/*?:"<>|]', '_', title)

    # --- Excel Reporting ---
    def _create_initial_report(self):
        excel_path = self.config.get('excel_report_path')
        if not excel_path: return
        try:
            self._send_log(f"Creando reporte inicial en: {excel_path}")
            df = self.config['input_df'].copy()
            df['Failure_Reason'] = 'En cola'

            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fallidos', index=False)
                pd.DataFrame(columns=list(df.columns) + ['Successful_Mirror']).to_excel(writer, sheet_name='Obtenidos', index=False)

            workbook = openpyxl.load_workbook(excel_path)
            if 'Fallidos' in workbook.sheetnames:
                workbook.move_sheet('Fallidos', offset=-len(workbook.sheetnames))
            workbook.save(excel_path)
        except Exception as e:
            self._send_log(f"Error al crear reporte Excel: {e}")

    def _update_report_on_success(self, successful_row_data, source):
        excel_path = self.config.get('excel_report_path')
        if not excel_path or not os.path.exists(excel_path): return

        try:
            doi_to_move = successful_row_data['DOI']
            xls = pd.ExcelFile(excel_path)
            df_fallidos = pd.read_excel(xls, sheet_name='Fallidos')
            df_obtenidos = pd.read_excel(xls, sheet_name='Obtenidos')
            xls.close()

            row_to_move_df = df_fallidos[df_fallidos['DOI'] == doi_to_move]
            df_fallidos = df_fallidos[df_fallidos['DOI'] != doi_to_move].copy()

            if not row_to_move_df.empty:
                row_to_move = row_to_move_df.iloc[0].to_dict()
                row_to_move['Successful_Mirror'] = source
                row_to_move.pop('Failure_Reason', None)

                new_obtenidos_row = pd.DataFrame([row_to_move])
                df_obtenidos = pd.concat([df_obtenidos, new_obtenidos_row], ignore_index=True)

            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_fallidos.to_excel(writer, sheet_name='Fallidos', index=False)
                df_obtenidos.to_excel(writer, sheet_name='Obtenidos', index=False)

            workbook = openpyxl.load_workbook(excel_path)
            if 'Fallidos' in workbook.sheetnames:
                workbook.move_sheet('Fallidos', offset=-len(workbook.sheetnames))
            workbook.save(excel_path)
        except Exception as e:
            self._send_log(f"Error al actualizar reporte Excel para DOI {doi_to_move}: {e}")

    def _update_report_on_failure(self, failed_row_data, reason):
        excel_path = self.config.get('excel_report_path')
        if not excel_path or not os.path.exists(excel_path): return

        try:
            doi_to_update = failed_row_data['DOI']
            xls = pd.ExcelFile(excel_path)
            df_fallidos = pd.read_excel(xls, sheet_name='Fallidos')
            df_obtenidos = pd.read_excel(xls, sheet_name='Obtenidos')
            xls.close()

            df_fallidos.loc[df_fallidos['DOI'] == doi_to_update, 'Failure_Reason'] = reason

            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_fallidos.to_excel(writer, sheet_name='Fallidos', index=False)
                df_obtenidos.to_excel(writer, sheet_name='Obtenidos', index=False)

            workbook = openpyxl.load_workbook(excel_path)
            if 'Fallidos' in workbook.sheetnames:
                workbook.move_sheet('Fallidos', offset=-len(workbook.sheetnames))
            workbook.save(excel_path)
        except Exception as e:
            self._send_log(f"Error al actualizar reporte de fallo para DOI {doi_to_update}: {e}")
