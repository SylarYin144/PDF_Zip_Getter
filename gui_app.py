import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import threading
import queue
import platform
import subprocess
import time
import json
from scihub_downloader import download_pdfs_from_file, DEFAULT_SCI_HUB_MIRRORS_EXAMPLE
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib import pyplot as plt
from collections import Counter
import openpyxl
from openpyxl.drawing.image import Image

class ConfigFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.input_file_path = ""
        self.zip_file_path = ""
        self.excel_report_path = ""
        self.grid_columnconfigure(0, weight=1); self.grid_columnconfigure(1, weight=1)
        io_frame = ctk.CTkFrame(self); io_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew"); io_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(io_frame, text="1. Archivos de Entrada y Salida", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10, 5), sticky="w")
        self.select_file_button = ctk.CTkButton(io_frame, text="Seleccionar Archivo (.xlsx, .csv)", command=self.select_input_file); self.select_file_button.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.file_info_label = ctk.CTkLabel(io_frame, text=""); self.file_info_label.grid(row=1, column=1, padx=10, pady=10, sticky="w")
        self.select_zip_button = ctk.CTkButton(io_frame, text="Definir Ubicación del ZIP", command=self.select_zip_location); self.select_zip_button.grid(row=2, column=0, padx=10, pady=10, sticky="w")
        self.zip_path_label = ctk.CTkLabel(io_frame, text=""); self.zip_path_label.grid(row=3, column=0, columnspan=2, padx=10, pady=(0, 10), sticky="w")
        sources_frame = ctk.CTkFrame(self); sources_frame.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")
        ctk.CTkLabel(sources_frame, text="2. Fuentes de Descarga", font=ctk.CTkFont(weight="bold")).pack(padx=10, pady=(10, 5), anchor="w")
        self.check_scihub = ctk.CTkCheckBox(sources_frame, text="Usar Sci-Hub"); self.check_scihub.pack(padx=10, pady=5, anchor="w")
        self.check_gscholar = ctk.CTkCheckBox(sources_frame, text="Usar Google Scholar"); self.check_gscholar.pack(padx=10, pady=5, anchor="w")
        self.check_pmc = ctk.CTkCheckBox(sources_frame, text="Usar PubMed Central (PMC)"); self.check_pmc.pack(padx=10, pady=(5, 10), anchor="w")
        adv_frame = ctk.CTkFrame(self); adv_frame.grid(row=0, column=1, rowspan=2, padx=20, pady=20, sticky="nsew")
        adv_frame.grid_columnconfigure(0, weight=1); adv_frame.grid_rowconfigure(2, weight=1)
        ctk.CTkLabel(adv_frame, text="3. Configuración Avanzada", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10, 5), sticky="w")
        ctk.CTkLabel(adv_frame, text="Mirrors de Sci-Hub (separados por coma):").grid(row=1, column=0, columnspan=2, padx=10, pady=(5,0), sticky="w")
        self.mirrors_textbox = ctk.CTkTextbox(adv_frame); self.mirrors_textbox.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="nsew")
        ctk.CTkLabel(adv_frame, text="Tiempo de espera entre DOIs (s):").grid(row=3, column=0, padx=10, pady=(10, 0), sticky="w")
        self.delay_entry = ctk.CTkEntry(adv_frame, width=80); self.delay_entry.grid(row=3, column=1, padx=10, pady=(10,0), sticky="e")
        ctk.CTkLabel(adv_frame, text="Timeout Carga de Página (s):").grid(row=4, column=0, padx=10, pady=(10, 0), sticky="w")
        self.page_load_timeout_entry = ctk.CTkEntry(adv_frame, width=80); self.page_load_timeout_entry.grid(row=4, column=1, padx=10, pady=(10,0), sticky="e")
        ctk.CTkLabel(adv_frame, text="Timeout Búsqueda Elemento (s):").grid(row=5, column=0, padx=10, pady=(10, 0), sticky="w")
        self.element_wait_timeout_entry = ctk.CTkEntry(adv_frame, width=80); self.element_wait_timeout_entry.grid(row=5, column=1, padx=10, pady=(10,0), sticky="e")
        report_frame = ctk.CTkFrame(self); report_frame.grid(row=2, column=0, padx=20, pady=20, sticky="w")
        self.report_button = ctk.CTkButton(report_frame, text="Definir Ruta de Reporte (.xlsx)", command=self.select_excel_report_path); self.report_button.pack(side="left", padx=(10,5), pady=10)
        self.report_path_label = ctk.CTkLabel(report_frame, text=""); self.report_path_label.pack(side="left", padx=5, pady=10)
        self.start_button = ctk.CTkButton(self, text="🚀 Iniciar Descarga", command=self.start_download, font=ctk.CTkFont(size=16)); self.start_button.grid(row=3, column=0, columnspan=2, padx=20, pady=20, sticky="ew")
        self.load_config()
        self._check_start_conditions()

    def load_config(self):
        try:
            with open("config.json", "r") as f:
                config = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            config = {}
        self.input_file_path = config.get("input_file_path", "")
        if self.input_file_path and os.path.exists(self.input_file_path):
            try:
                df = pd.read_csv(self.input_file_path, dtype=str) if self.input_file_path.endswith('.csv') else pd.read_excel(self.input_file_path, dtype=str)
                df.fillna('', inplace=True); df.rename(columns={col: col.lower() for col in df.columns}, inplace=True)
                if 'doi' not in df.columns: raise ValueError("El archivo no contiene la columna 'DOI'.")
                self.file_info_label.configure(text=f"{os.path.basename(self.input_file_path)} ({len(df)} artículos detectados)")
                self.controller.input_df = df
            except Exception as e:
                self.file_info_label.configure(text=f"Ruta guardada, pero error al cargar: {e}", text_color="orange")
                self.input_file_path = ""; self.controller.input_df = None
        else:
            self.file_info_label.configure(text="No se ha cargado ningún archivo.")
        self.zip_file_path = config.get("zip_file_path", "")
        self.zip_path_label.configure(text=f"Se guardará en: {self.zip_file_path}" if self.zip_file_path else "No se ha seleccionado la ubicación.")
        self.excel_report_path = config.get("excel_report_path", "")
        self.report_path_label.configure(text=f"Reporte se guardará en: {self.excel_report_path}" if self.excel_report_path else "No se generará reporte.")
        self.check_scihub.set(config.get("use_scihub", 1))
        self.check_gscholar.set(config.get("use_gscholar", 1))
        self.check_pmc.set(config.get("use_pmc", 1))
        self.mirrors_textbox.delete("1.0", "end")
        self.mirrors_textbox.insert("1.0", config.get("mirrors", ",\n".join(DEFAULT_SCI_HUB_MIRRORS_EXAMPLE)))
        self.delay_entry.delete(0, "end"); self.delay_entry.insert(0, str(config.get("delay", "2")))
        self.page_load_timeout_entry.delete(0, "end"); self.page_load_timeout_entry.insert(0, str(config.get("page_load_timeout", "60")))
        self.element_wait_timeout_entry.delete(0, "end"); self.element_wait_timeout_entry.insert(0, str(config.get("element_wait_timeout", "20")))

    def save_config(self):
        config = {"input_file_path": self.input_file_path, "zip_file_path": self.zip_file_path, "excel_report_path": self.excel_report_path, "use_scihub": self.check_scihub.get(), "use_gscholar": self.check_gscholar.get(), "use_pmc": self.check_pmc.get(), "mirrors": self.mirrors_textbox.get("1.0", "end-1c"), "delay": self.delay_entry.get(), "page_load_timeout": self.page_load_timeout_entry.get(), "element_wait_timeout": self.element_wait_timeout_entry.get()}
        try:
            with open("config.json", "w") as f:
                json.dump(config, f, indent=4)
        except Exception as e:
            print(f"Advertencia: No se pudo guardar el archivo de configuración. {e}")

    def select_input_file(self):
        path = filedialog.askopenfilename(title="Seleccionar archivo con DOIs", filetypes=(("Archivos Excel", "*.xlsx *.xls"), ("Archivos CSV", "*.csv")))
        if not path: return
        self.input_file_path = path
        try:
            df = pd.read_csv(path, dtype=str) if path.endswith('.csv') else pd.read_excel(path, dtype=str)
            df.fillna('', inplace=True); df.rename(columns={col: col.lower() for col in df.columns}, inplace=True)
            if 'doi' not in df.columns: raise ValueError("El archivo no contiene la columna 'DOI'.")
            self.file_info_label.configure(text=f"{os.path.basename(path)} ({len(df)} artículos detectados)")
            self.controller.input_df = df
        except Exception as e: self.file_info_label.configure(text=f"Error al leer archivo: {e}", text_color="red"); self.input_file_path = ""
        self._check_start_conditions()

    def select_zip_location(self):
        path = filedialog.asksaveasfilename(title="Guardar archivo ZIP como...", defaultextension=".zip", filetypes=(("Archivos ZIP", "*.zip"),))
        if not path: return
        self.zip_file_path = path; self.zip_path_label.configure(text=f"Se guardará en: {path}"); self._check_start_conditions()

    def select_excel_report_path(self):
        path = filedialog.asksaveasfilename(title="Guardar Reporte Excel como...", defaultextension=".xlsx", filetypes=(("Archivos Excel", "*.xlsx"),))
        if not path: self.excel_report_path = ""; self.report_path_label.configure(text="No se generará reporte.")
        else: self.excel_report_path = path; self.report_path_label.configure(text=f"Reporte se guardará en: {path}")

    def _check_start_conditions(self):
        if self.input_file_path and self.zip_file_path and self.controller.input_df is not None: self.start_button.configure(state="normal")
        else: self.start_button.configure(state="disabled")

    def start_download(self):
        self.save_config()
        if not self.check_scihub.get() and not self.check_gscholar.get() and not self.check_pmc.get(): self.controller.show_error("Sin fuentes", "Por favor, seleccione al menos una fuente de descarga."); return
        mirrors = [m.strip() for m in self.mirrors_textbox.get("1.0", "end-1c").split(',') if m.strip()] if self.check_scihub.get() else []
        if self.check_scihub.get() and not mirrors: self.controller.show_error("Mirrors de Sci-Hub vacíos", "Por favor, especifique al menos un mirror de Sci-Hub si la fuente está activada."); return
        try: delay = int(self.delay_entry.get()); page_load_timeout = int(self.page_load_timeout_entry.get()); element_wait_timeout = int(self.element_wait_timeout_entry.get())
        except ValueError: self.controller.show_error("Valor inválido", "Los tiempos de espera deben ser números enteros."); return
        config = {'df': self.controller.input_df, 'zip_path': self.zip_file_path, 'excel_report_path': self.excel_report_path, 'sci_hub_mirrors': mirrors, 'use_google_scholar': bool(self.check_gscholar.get()), 'use_pmc': bool(self.check_pmc.get()), 'inter_doi_delay': delay, 'page_load_timeout': page_load_timeout, 'element_wait_timeout': element_wait_timeout}
        self.controller.start_download_process(config)

class ProgressFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.grid_columnconfigure(1, weight=1); self.grid_rowconfigure(1, weight=1)
        top_frame = ctk.CTkFrame(self, fg_color="transparent"); top_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=5)
        top_frame.grid_columnconfigure(0, weight=1); top_frame.grid_columnconfigure(1, weight=1)
        self.current_article_frame = ctk.CTkFrame(top_frame); self.current_article_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        self.current_article_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(self.current_article_frame, text="Procesando Artículo:", font=ctk.CTkFont(size=14, weight="bold")).pack(padx=10, pady=(5,0), anchor="w")
        self.current_article_num_label = ctk.CTkLabel(self.current_article_frame, text="Artículo: --/--", anchor="w"); self.current_article_num_label.pack(padx=10, pady=2, anchor="w")
        self.current_title_label = ctk.CTkLabel(self.current_article_frame, text="Título: --", anchor="w", wraplength=400, justify="left"); self.current_title_label.pack(padx=10, pady=2, anchor="w")
        self.current_author_label = ctk.CTkLabel(self.current_article_frame, text="Autor: --", anchor="w"); self.current_author_label.pack(padx=10, pady=2, anchor="w")
        self.current_journal_label = ctk.CTkLabel(self.current_article_frame, text="Revista: --", anchor="w"); self.current_journal_label.pack(padx=10, pady=2, anchor="w")
        self.current_year_label = ctk.CTkLabel(self.current_article_frame, text="Año: --", anchor="w"); self.current_year_label.pack(padx=10, pady=2, anchor="w")
        self.current_doi_label = ctk.CTkLabel(self.current_article_frame, text="DOI: --", anchor="w"); self.current_doi_label.pack(padx=10, pady=(2,5), anchor="w")
        time_kpi_frame = ctk.CTkFrame(top_frame); time_kpi_frame.grid(row=0, column=1, sticky="nsew", padx=(5,0))
        time_kpi_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(time_kpi_frame, text="Métricas de Tiempo", font=ctk.CTkFont(weight="bold")).pack(pady=5)
        self.time_elapsed_label = ctk.CTkLabel(time_kpi_frame, text="Transcurrido: 00:00:00"); self.time_elapsed_label.pack(pady=2)
        self.time_avg_label = ctk.CTkLabel(time_kpi_frame, text="Promedio/DOI: -- s"); self.time_avg_label.pack(pady=2)
        self.time_etr_label = ctk.CTkLabel(time_kpi_frame, text="Tiempo Restante: --:--:--"); self.time_etr_label.pack(pady=(2,5))
        main_content_frame = ctk.CTkFrame(self, fg_color="transparent"); main_content_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=10, pady=5)
        main_content_frame.grid_columnconfigure(0, weight=1); main_content_frame.grid_columnconfigure(1, weight=1); main_content_frame.grid_rowconfigure(0, weight=1)
        self.article_list_frame = ctk.CTkScrollableFrame(main_content_frame, label_text="Lista de Artículos"); self.article_list_frame.grid(row=0, column=0, sticky="nsew", padx=(0,5))
        self.article_widgets = {}
        charts_parent_frame = ctk.CTkFrame(main_content_frame, fg_color="transparent"); charts_parent_frame.grid(row=0, column=1, sticky="nsew", padx=(5,0))
        charts_parent_frame.grid_columnconfigure(0, weight=1); charts_parent_frame.grid_rowconfigure((0,1,2), weight=1)
        self.chart1_frame = ctk.CTkFrame(charts_parent_frame); self.chart1_frame.grid(row=0, column=0, sticky="nsew", pady=(0,5))
        self.chart2_frame = ctk.CTkFrame(charts_parent_frame); self.chart2_frame.grid(row=1, column=0, sticky="nsew", pady=5)
        self.chart3_frame = ctk.CTkFrame(charts_parent_frame); self.chart3_frame.grid(row=2, column=0, sticky="nsew", pady=(5,0))
        self.chart1_canvas = self.chart2_canvas = self.chart3_canvas = None
        log_frame = ctk.CTkFrame(self); log_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=5)
        log_frame.grid_columnconfigure(0, weight=1)
        self.log_textbox = ctk.CTkTextbox(log_frame, height=120); self.log_textbox.grid(row=0, column=0, sticky="ew"); self.log_textbox.configure(state="disabled")
        bottom_frame = ctk.CTkFrame(self); bottom_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=10, pady=10)
        bottom_frame.grid_columnconfigure(0, weight=1)
        self.progress_bar = ctk.CTkProgressBar(bottom_frame); self.progress_bar.grid(row=0, column=0, padx=10, pady=5, sticky="ew"); self.progress_bar.set(0)
        self.progress_label = ctk.CTkLabel(bottom_frame, text="Procesando 0 de 0... (0%)"); self.progress_label.grid(row=1, column=0, padx=10, pady=(0, 5))
        controls_frame = ctk.CTkFrame(bottom_frame, fg_color="transparent"); controls_frame.grid(row=2, column=0, sticky="ew", pady=5)
        controls_frame.grid_columnconfigure((0,1), weight=1)
        self.pause_button = ctk.CTkButton(controls_frame, text="Pausar", command=self.controller.toggle_pause); self.pause_button.grid(row=0, column=0, padx=5)
        self.cancel_button = ctk.CTkButton(controls_frame, text="Cancelar", fg_color="red", hover_color="darkred", command=self.controller.cancel_download); self.cancel_button.grid(row=0, column=1, padx=5)
    def reset_ui(self, total_articles):
        self.current_article_num_label.configure(text=f"Artículo: 0/{total_articles} (0.0%)"); self.current_title_label.configure(text="Título: Esperando para iniciar..."); self.current_author_label.configure(text="Autor: --"); self.current_journal_label.configure(text="Revista: --"); self.current_year_label.configure(text="Año: --"); self.current_doi_label.configure(text="DOI: --")
        self.time_elapsed_label.configure(text="Transcurrido: 00:00:00"); self.time_avg_label.configure(text="Promedio/DOI: -- s"); self.time_etr_label.configure(text="Tiempo Restante: --:--:--")
        self.log_textbox.configure(state="normal"); self.log_textbox.delete("1.0", "end"); self.log_textbox.configure(state="disabled")
        for widget in self.article_list_frame.winfo_children(): widget.destroy()
        self.article_widgets = {}; self.pause_button.configure(text="Pausar", state="normal"); self.cancel_button.configure(state="normal")
        for canvas in [self.chart1_canvas, self.chart2_canvas, self.chart3_canvas]:
            if canvas: canvas.get_tk_widget().destroy()
        for frame in [self.chart1_frame, self.chart2_frame, self.chart3_frame]:
            for widget in frame.winfo_children(): widget.destroy()
        self.chart1_canvas = self.chart2_canvas = self.chart3_canvas = None
    def populate_initial_articles(self, df):
        self.article_list_frame.grid_columnconfigure(1, weight=1)
        for index, row in df.iterrows():
            doi = row.get("doi", "N/A"); title = row.get("title", doi) if row.get("title") else doi
            journal = row.get('journal/book', row.get('revista', '')); year = row.get('publication year', row.get('año', ''))
            extra_info = f" ({journal}, {year})" if journal and year else ""
            status_label = ctk.CTkLabel(self.article_list_frame, text="⏳ En cola", width=120); status_label.grid(row=index, column=0, padx=5, pady=2, sticky="w")
            title_label = ctk.CTkLabel(self.article_list_frame, text=f"{title}{extra_info}", anchor="w", wraplength=400, justify="left"); title_label.grid(row=index, column=1, padx=5, pady=2, sticky="ew")
            self.article_widgets[doi] = {'status': status_label, 'title': title_label}
    def update_current_article(self, article_data):
        self.current_article_num_label.configure(text=f"Artículo: {article_data['current']}/{article_data['total']} ({article_data['current']/article_data['total']*100:.1f}%)")
        self.current_title_label.configure(text=f"Título: {article_data.get('title', 'N/A')}")
        self.current_author_label.configure(text=f"Autor: {article_data.get('first author', 'N/A')}")
        self.current_journal_label.configure(text=f"Revista: {article_data.get('journal/book', article_data.get('revista', 'N/A'))}")
        self.current_year_label.configure(text=f"Año: {article_data.get('publication year', article_data.get('año', 'N/A'))}")
        self.current_doi_label.configure(text=f"DOI: {article_data.get('doi', 'N/A')}")
    def update_timers(self, elapsed, processed_count, pending_count):
        self.time_elapsed_label.configure(text=f"Transcurrido: {time.strftime('%H:%M:%S', time.gmtime(elapsed))}")
        avg = elapsed / processed_count if processed_count > 0 else 0
        etr = avg * pending_count if avg > 0 else 0
        self.time_avg_label.configure(text=f"Promedio/DOI: {avg:.2f} s" if avg > 0 else "Promedio/DOI: -- s")
        self.time_etr_label.configure(text=f"Tiempo Restante: {time.strftime('%H:%M:%S', time.gmtime(etr))}" if etr > 0 else "Tiempo Restante: --:--:--")
    def update_charts(self, kpi_data):
        self.chart1_canvas = self._create_chart(self.chart1_frame, self.chart1_canvas, {'title': 'Progreso General', 'labels': ['Obtenidos', 'Fallidos', 'Pendientes'], 'sizes': [kpi_data['obtained'], kpi_data['failed'], kpi_data['pending']], 'colors': ['#34A853', '#EA4335', 'grey']})
        self.chart2_canvas = self._create_chart(self.chart2_frame, self.chart2_canvas, {'title': 'Tasa de Éxito (de Procesados)', 'labels': ['Obtenidos', 'Fallidos'], 'sizes': [kpi_data['obtained'], kpi_data['failed']], 'colors': ['#34A853', '#EA4335']})
        source_counts = kpi_data['source_counts'].copy()
        if kpi_data['failed'] > 0: source_counts['Fallidos'] = kpi_data['failed']
        chart3_data = {'title': 'Desglose de Resultados', 'labels': list(source_counts.keys()), 'sizes': list(source_counts.values()), 'wedgeprops': dict(width=0.4)}
        self.chart3_canvas = self._create_chart(self.chart3_frame, self.chart3_canvas, chart3_data)
    def _create_chart(self, frame, canvas, chart_data):
        if canvas: canvas.get_tk_widget().destroy()
        for widget in frame.winfo_children(): widget.destroy()
        if not chart_data['sizes'] or sum(chart_data['sizes']) == 0: ctk.CTkLabel(frame, text=f"{chart_data['title']}\n(Esperando datos)").pack(expand=True); return
        is_dark = ctk.get_appearance_mode() == "Dark"; bg_color = "#242424" if is_dark else "#dbdbdb"; text_color = "#FFFFFF" if is_dark else "#000000"
        fig = Figure(figsize=(4, 2.5), dpi=80, facecolor=bg_color)
        ax = fig.add_subplot(111)
        wedges, texts, autotexts = ax.pie(chart_data['sizes'], autopct='%1.1f%%', startangle=90, colors=chart_data.get('colors'), wedgeprops=chart_data.get('wedgeprops', {}), textprops={'color': text_color, 'fontsize': 8})
        ax.legend(wedges, chart_data['labels'], title="Categorías", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), prop={'size': 8}, title_fontproperties={'size':9, 'weight':'bold', 'color':text_color}, labelcolor=text_color)
        ax.set_title(chart_data['title'], color=text_color, fontsize=10); fig.tight_layout(pad=1.5)
        new_canvas = FigureCanvasTkAgg(fig, master=frame); new_canvas.draw(); new_canvas.get_tk_widget().pack(side=ctk.TOP, fill=ctk.BOTH, expand=True)
        return new_canvas
    def update_article_status(self, doi, success, source, reason):
        if doi in self.article_widgets:
            status_text = f"✅ Obtenido ({source})" if success else f"❌ Fallido"; color = "#34A853" if success else "#EA4335"
            self.article_widgets[doi]['status'].configure(text=status_text, text_color=color)
    def add_log_message(self, message):
        self.log_textbox.configure(state="normal"); self.log_textbox.insert("end", message + "\n"); self.log_textbox.see("end"); self.log_textbox.configure(state="disabled")
    def finalize_ui(self):
        self.pause_button.configure(state="disabled"); self.cancel_button.configure(state="disabled")

class ResultsFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller; self.results_data = {}; self.grid_columnconfigure(0, weight=1); self.grid_rowconfigure(2, weight=1)
        self.summary_label = ctk.CTkLabel(self, text="", font=ctk.CTkFont(size=20, weight="bold")); self.summary_label.grid(row=0, column=0, padx=20, pady=10)
        charts_parent_frame = ctk.CTkFrame(self, fg_color="transparent"); charts_parent_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        charts_parent_frame.grid_columnconfigure((0,1,2), weight=1)
        self.chart1_frame = ctk.CTkFrame(charts_parent_frame); self.chart1_frame.grid(row=0, column=0, sticky="nsew", padx=5)
        self.chart2_frame = ctk.CTkFrame(charts_parent_frame); self.chart2_frame.grid(row=0, column=1, sticky="nsew", padx=5)
        self.chart3_frame = ctk.CTkFrame(charts_parent_frame); self.chart3_frame.grid(row=0, column=2, sticky="nsew", padx=5)
        self.chart1_canvas = self.chart2_canvas = self.chart3_canvas = None
        self.tab_view = ctk.CTkTabview(self); self.tab_view.grid(row=2, column=0, padx=20, pady=10, sticky="nsew")
        self.tab_view.add("Obtenidos"); self.tab_view.add("Fallidos")
        self.obtained_frame = ctk.CTkScrollableFrame(self.tab_view.tab("Obtenidos")); self.obtained_frame.pack(fill="both", expand=True)
        self.failed_frame = ctk.CTkScrollableFrame(self.tab_view.tab("Fallidos")); self.failed_frame.pack(fill="both", expand=True)
        self.copy_failed_button = ctk.CTkButton(self.tab_view.tab("Fallidos"), text="Copiar DOIs fallidos", command=self.copy_failed_dois); self.copy_failed_button.pack(pady=10)
        actions_frame = ctk.CTkFrame(self); actions_frame.grid(row=3, column=0, padx=20, pady=20, sticky="ew")
        actions_frame.grid_columnconfigure((0,1,2), weight=1)
        self.open_folder_button = ctk.CTkButton(actions_frame, text="📂 Abrir Carpeta de Descarga", command=self.open_download_folder); self.open_folder_button.grid(row=0, column=0, padx=5, pady=10)
        self.open_report_button = ctk.CTkButton(actions_frame, text="🧾 Ver Reporte en Excel", command=self.open_excel_report); self.open_report_button.grid(row=0, column=1, padx=5, pady=10)
        self.new_task_button = ctk.CTkButton(actions_frame, text="🔄 Iniciar una Nueva Tarea", command=lambda: controller.show_frame("ConfigFrame")); self.new_task_button.grid(row=0, column=2, padx=5, pady=10)
    def update_results(self, results):
        self.results_data = results; total = results['total_articles']; success = results['successful_count']
        msg = "¡Proceso Completado!" if not results['was_cancelled'] else "Proceso Cancelado"
        self.summary_label.configure(text=f"{msg} Se obtuvieron {success} de {total} artículos.")
        for widget in self.obtained_frame.winfo_children(): widget.destroy()
        for item in results['successful_articles']: ctk.CTkLabel(self.obtained_frame, text=f"✅ {item['data'].get('title', item['data']['doi'])} (Fuente: {item['source']})").pack(anchor="w", padx=5)
        for widget in self.failed_frame.winfo_children(): widget.destroy()
        for item in results['failed_articles']: ctk.CTkLabel(self.failed_frame, text=f"❌ {item['data'].get('title', item['data']['doi'])} (Razón: {item['reason']})").pack(anchor="w", padx=5)
        self.copy_failed_button.pack_forget() if not results['failed_articles'] else self.copy_failed_button.pack(pady=10)
        self.open_report_button.configure(state="normal" if results['excel_report_path'] else "disabled")
        self.update_final_charts(results)
        if self.results_data.get('excel_report_path'): self.export_charts_to_excel(results)
    def update_final_charts(self, results):
        pending_count = results['total_articles'] - results['successful_count'] - results['failed_count']
        chart1_data = {'title': 'Progreso General', 'labels': ['Obtenidos', 'Fallidos', 'Pendientes'], 'sizes': [results['successful_count'], results['failed_count'], pending_count], 'colors': ['#34A853', '#EA4335', 'grey']}
        self.chart1_canvas = self._create_chart(self.chart1_frame, self.chart1_canvas, chart1_data)
        chart2_data = {'title': 'Tasa de Éxito (de Procesados)', 'labels': ['Obtenidos', 'Fallidos'], 'sizes': [results['successful_count'], results['failed_count']], 'colors': ['#34A853', '#EA4335']}
        self.chart2_canvas = self._create_chart(self.chart2_frame, self.chart2_canvas, chart2_data)
        source_counts = results.get('source_counts', Counter())
        if results['failed_count'] > 0: source_counts['Fallidos'] = results['failed_count']
        chart3_data = {'title': 'Desglose de Resultados', 'labels': list(source_counts.keys()), 'sizes': list(source_counts.values()), 'wedgeprops': dict(width=0.4)}
        self.chart3_canvas = self._create_chart(self.chart3_frame, self.chart3_canvas, chart3_data)
    def _create_chart(self, frame, canvas, chart_data):
        if canvas: canvas.get_tk_widget().destroy()
        for widget in frame.winfo_children(): widget.destroy()
        if not chart_data['sizes'] or sum(chart_data['sizes']) == 0: ctk.CTkLabel(frame, text=f"{chart_data['title']}\n(No hay datos)").pack(expand=True); return
        is_dark = ctk.get_appearance_mode() == "Dark"; bg_color = "#242424" if is_dark else "#dbdbdb"; text_color = "#FFFFFF" if is_dark else "#000000"
        fig = Figure(figsize=(4, 2.5), dpi=100, facecolor=bg_color)
        ax = fig.add_subplot(111)
        wedges, texts, autotexts = ax.pie(chart_data['sizes'], autopct='%1.1f%%', startangle=90, colors=chart_data.get('colors'), wedgeprops=chart_data.get('wedgeprops', {}), textprops={'color': text_color, 'fontsize': 8})
        ax.legend(wedges, chart_data['labels'], title="Categorías", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), prop={'size': 8}, title_fontproperties={'size':9, 'weight':'bold', 'color':text_color}, labelcolor=text_color)
        ax.set_title(chart_data['title'], color=text_color, fontsize=10); fig.tight_layout(pad=1.5)
        new_canvas = FigureCanvasTkAgg(fig, master=frame); new_canvas.draw(); new_canvas.get_tk_widget().pack(side=ctk.TOP, fill=ctk.BOTH, expand=True)
        return new_canvas
    def export_charts_to_excel(self, results):
        excel_path = results.get('excel_report_path')
        self.controller.add_log_message("Exportando gráficos a Excel...")
        chart_paths = []
        try:
            fig1 = self._create_figure_for_export({'title': 'Progreso General', 'labels': ['Obtenidos', 'Fallidos', 'Pendientes'], 'sizes': [results['successful_count'], results['failed_count'], results['total_articles'] - results['successful_count'] - results['failed_count']], 'colors': ['#34A853', '#EA4335', 'grey']})
            if fig1: chart1_path = "chart_progreso.png"; fig1.savefig(chart1_path); plt.close(fig1); chart_paths.append(chart1_path)
            fig2 = self._create_figure_for_export({'title': 'Tasa de Éxito (de Procesados)', 'labels': ['Obtenidos', 'Fallidos'], 'sizes': [results['successful_count'], results['failed_count']], 'colors': ['#34A853', '#EA4335']})
            if fig2: chart2_path = "chart_tasa_exito.png"; fig2.savefig(chart2_path); plt.close(fig2); chart_paths.append(chart2_path)
            source_counts = results.get('source_counts', Counter());
            if results['failed_count'] > 0: source_counts['Fallidos'] = results['failed_count']
            fig3 = self._create_figure_for_export({'title': 'Desglose de Resultados', 'labels': list(source_counts.keys()), 'sizes': list(source_counts.values()), 'wedgeprops': dict(width=0.4)})
            if fig3: chart3_path = "chart_desglose.png"; fig3.savefig(chart3_path); plt.close(fig3); chart_paths.append(chart3_path)
            if not chart_paths: self.controller.add_log_message("No hay datos suficientes para generar gráficos."); return
            workbook = openpyxl.load_workbook(excel_path)
            if "Gráficos" not in workbook.sheetnames: charts_sheet = workbook.create_sheet("Gráficos", 0)
            else: charts_sheet = workbook["Gráficos"]
            for i, path in enumerate(chart_paths):
                img = Image(path); charts_sheet.add_image(img, f"{chr(ord('A') + i*9)}1")
            workbook.save(excel_path)
            self.controller.add_log_message("Gráficos exportados a Excel correctamente.")
        except Exception as e: self.controller.add_log_message(f"Error al exportar gráficos a Excel: {e}")
        finally:
            for path in chart_paths:
                if os.path.exists(path): os.remove(path)
    def _create_figure_for_export(self, chart_data):
        if not chart_data['sizes'] or sum(chart_data['sizes']) == 0: return None
        fig, ax = plt.subplots(figsize=(6, 4), dpi=100)
        ax.pie(chart_data['sizes'], labels=chart_data['labels'], autopct='%1.1f%%', startangle=90, colors=chart_data.get('colors'), wedgeprops=chart_data.get('wedgeprops', {}))
        ax.set_title(chart_data['title']); ax.axis('equal'); return fig
    def copy_failed_dois(self):
        failed_dois = [item['data']['doi'] for item in self.results_data.get('failed_articles', []) if item['data']['doi']]
        if failed_dois: self.clipboard_clear(); self.clipboard_append("\n".join(failed_dois)); self.controller.show_info("Copiado", f"{len(failed_dois)} DOIs fallidos copiados al portapapeles.")
    def open_path(self, path):
        if not path or not os.path.exists(path): self.controller.show_error("Error", f"La ruta no existe o no fue especificada:\n{path}"); return
        try:
            if platform.system() == "Windows": os.startfile(os.path.dirname(path) if os.path.isfile(path) else path)
            elif platform.system() == "Darwin": subprocess.run(["open", path if os.path.isdir(path) else os.path.dirname(path)])
            else: subprocess.run(["xdg-open", path if os.path.isdir(path) else os.path.dirname(path)])
        except Exception as e: self.controller.show_error("Error al abrir", f"No se pudo abrir la ruta:\n{e}")
    def open_download_folder(self): self.open_path(self.results_data.get('zip_path'))
    def open_excel_report(self): self.open_path(self.results_data.get('excel_report_path'))

class App(ctk.CTk):
    def __init__(self):
        super().__init__(); self.title("Sci-Hub Downloader Pro"); self.geometry("1200x900"); self.input_df = None; self.downloader_thread = None
        self.progress_queue = queue.Queue(); self.cancel_event = threading.Event(); self.pause_event = threading.Event(); self.start_time = 0; self.timer_id = None; self.kpi_data = {}
        ctk.set_appearance_mode("System"); ctk.set_default_color_theme("blue")
        container = ctk.CTkFrame(self); container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1); container.grid_columnconfigure(0, weight=1)
        self.frames = {}
        for F in (ConfigFrame, ProgressFrame, ResultsFrame):
            frame = F(parent=container, controller=self); self.frames[F.__name__] = frame; frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame("ConfigFrame")

    def show_frame(self, frame_name): self.frames[frame_name].tkraise()
    def start_download_process(self, config):
        self.cancel_event.clear(); self.pause_event.clear(); self.kpi_data = {'obtained': 0, 'failed': 0, 'pending': len(config['df']), 'source_counts': Counter()}
        progress_frame = self.frames['ProgressFrame']; progress_frame.reset_ui(len(config['df'])); progress_frame.populate_initial_articles(config['df'])
        self.show_frame("ProgressFrame")
        self.start_time = time.time()
        self.downloader_thread = threading.Thread(target=download_pdfs_from_file, args=(config, self.progress_queue, self.cancel_event, self.pause_event))
        self.downloader_thread.start()
        self.after(100, self.process_queue)
        self.update_time_kpis()
    def process_queue(self):
        try:
            message = self.progress_queue.get_nowait()
            progress_frame = self.frames['ProgressFrame']
            if message['type'] == 'log': progress_frame.add_log_message(message['message'])
            elif message['type'] == 'current_article': progress_frame.update_current_article(message['data'])
            elif message['type'] == 'kpi': self.kpi_data = message; progress_frame.update_charts(self.kpi_data)
            elif message['type'] == 'article_result': progress_frame.update_article_status(message['doi'], message['success'], message.get('source'), message.get('reason'))
            elif message['type'] == 'finished':
                if self.timer_id: self.after_cancel(self.timer_id)
                progress_frame.finalize_ui()
                self.frames['ResultsFrame'].update_results(message['summary'])
                self.show_frame("ResultsFrame")
                return
        except queue.Empty: pass
        finally:
            if self.downloader_thread and self.downloader_thread.is_alive(): self.after(100, self.process_queue)
            else:
                while not self.progress_queue.empty():
                    try:
                        message = self.progress_queue.get_nowait()
                        if message['type'] == 'finished': self.frames['ResultsFrame'].update_results(message['summary']); self.show_frame("ResultsFrame")
                    except queue.Empty: break
    def update_time_kpis(self):
        if not self.downloader_thread or not self.downloader_thread.is_alive():
            if self.timer_id: self.after_cancel(self.timer_id); self.timer_id = None
            return
        elapsed = time.time() - self.start_time
        processed_count = self.kpi_data.get('obtained', 0) + self.kpi_data.get('failed', 0)
        pending_count = self.kpi_data.get('pending', 0)
        self.frames['ProgressFrame'].update_timers(elapsed, processed_count, pending_count)
        self.timer_id = self.after(1000, self.update_time_kpis)
    def toggle_pause(self):
        if self.pause_event.is_set(): self.pause_event.clear(); self.frames['ProgressFrame'].pause_button.configure(text="Pausar")
        else: self.pause_event.set(); self.frames['ProgressFrame'].pause_button.configure(text="Reanudar")
    def cancel_download(self): self.cancel_event.set(); self.frames['ProgressFrame'].finalize_ui()
    def add_log_message(self, msg): self.frames['ProgressFrame'].add_log_message(msg)
    def show_error(self, title, message): messagebox.showerror(title, message)
    def show_info(self, title, message): messagebox.showinfo(title, message)

if __name__ == "__main__":
    app = App()
    app.mainloop()
